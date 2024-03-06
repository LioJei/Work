import os
import random
import sys
import shelve
import openpyxl
import time
import datetime


class TestDemo:
    """
    @brief:           测试类初始化函数
    @param[in]:       name(用例名)
    """

    def __init__(self, name):
        self.__name = str(name)

    """
    @brief:           打印函数
    @param[in]:       msg(要打印的信息)
    """

    def test_print(self, msg):
        print(self.__name + ":" + str(msg))

    """
    @brief:           输入测试函数
    @param[in]:       m_num(匹配数)
    """

    def input_test(self, m_num):
        while True:
            num = int(input(self.__name + ":please guess a number:"))
            if num == int(m_num):
                break
            else:
                self.test_print("try again!")

        self.test_print("ok!")

    """
    @brief:           random测试
    @param[in]:       m_num(测试次数)
    """

    def random_test(self, m_num):
        for i in range(m_num):
            self.test_print(random.randint(1, 10))

        self.test_print("finish random test.")

    """
    @brief:           sys_exit测试
    @param[in]:       m_str(匹配字串)
    """

    def sys_exit_test(self, m_str):
        while True:
            self.test_print("Type str to exit:")
            res = input()
            if res == m_str:
                sys.exit("sys_exit!test end.")

            self.test_print("please try again.")

    """
    @brief:           异常处理测试
    @param[in]:       m_num(被除数)
    """

    def div_catch_test(self, m_num):
        try:
            div_num = int(input("please enter a number:"))
            return m_num / div_num
        except ZeroDivisionError:
            self.test_print("Error! please check your number.")

    """
    @brief:           文件读写测试
    @param[in]:       file_name(文件名)
    @param[in]:       mode(文件格式)
    """

    def file_read_write_test(self, file_name, mode):
        os.mkdir('test_file')
        if mode == 'f':
            test_file = open(file_name, "a+", encoding="utf-8")
            for line in test_file.readlines():
                self.test_print(line)
            test_file.write('ha,ha,ha,ha,ha\n')
            test_file.close()
        else:
            test_file = shelve.open(file_name)
            self.test_print(type(test_file))
            test_data = ['hello', 'world', '!!!']
            test_file['test_data'] = test_data
            test_file.close()
            test_file = shelve.open(file_name)
            self.test_print(list(test_file.keys()))
            self.test_print(list(test_file.values()))
            test_file.close()

    """
    @brief:         算术题生成函数
    @param[in]:     limit_num(最大数)
    """

    def calculate_test(self, limit_num):
        file = open(os.getcwd() + '\\test_file\\work.txt', 'a+')
        simple = ['+', '-', 'x', '÷']
        for i in range(limit_num):
            num1 = random.randint(1, 100)
            num2 = random.randint(1, num1)
            sim = random.choice(simple)
            file.write(str(num1) + str(sim) + str(num2) + '=\n')

        file.flush()
        self.test_print('finish')
        file.close()

    """
    @brief:         Excel表格读写测试
    @param[in]:     
    """

    def excel_test(self, mode):
        if mode == 'r':
            wb = openpyxl.load_workbook(os.getcwd() +
                                        '\\test_file\\test.xlsx')
            self.test_print(type(wb))
            self.test_print(wb.sheetnames)
            sheet = wb.get_sheet_by_name('Sheet1')
            self.test_print(sheet)
            self.test_print(type(sheet))
            self.test_print(sheet.title)
            self.test_print(sheet.max_row)
            self.test_print(sheet.max_column)
            m_cell = sheet['A1']
            self.test_print(m_cell.value)
            self.test_print('Row ' + str(m_cell.row) + ', Column ' +
                            str(m_cell.column) + ' is ' + m_cell.value)
            self.test_print('Cell ' + m_cell.coordinate
                            + ' is ' + m_cell.value)
        elif mode == 'c':
            wb = openpyxl.load_workbook(os.getcwd() +
                                        '\\test_file\\test.xlsx')
            self.test_print(wb.get_sheet_names())
            wb.create_sheet(index=3, title='creat_sheet')
            self.test_print(wb.sheetnames)
            wb.remove_sheet(wb.get_sheet_by_name('Sheet3'))
            self.test_print(wb.sheetnames)
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.cell(row=1, column=2).value = 'world'
            sheet.merge_cells('A1:B1')
            sheet = wb.get_sheet_by_name('creat_sheet')
            for i in range(1, 10):
                sheet.cell(row=i, column=1).value = random.randint(1, 100)
            self.test_print(sheet['A9'].value)
            sheet.freeze_panes = 'A2'
            sheet['A10'].value = '=SUM(A1:A9)'
            wb.save(os.getcwd() + '\\test_file\\c_test_sheet.xlsx')

        else:
            wb = openpyxl.load_workbook(os.getcwd() +
                                        '\\test_file\\test.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.title = "write_test_sheet"
            wb.save(os.getcwd() + '\\test_file\\w_test_sheet.xlsx')

    """
    @brief:         时间模块测试
    """
    def time_test(self):
        self.test_print(time.time())
        self.test_print(datetime.datetime.now().
                        strftime('%Y-%m-%d %H:%M:%S %A'))
